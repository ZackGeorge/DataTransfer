import openpyxl


class DataTransfer:

    def __init__(self,):
        pass



    #sheet.cell(column = 3, row = 5).value = sheetF.cell(column = 3, row = 5).value

    def transfer(self, obs, wbT, sT, wbF, sF, columnHits, columnHitsToTake, rowHits, fileName):
        print "Observer: ", obs
        observer = obs
        columnTPHits = columnHits
        columnTPFA = columnHits+1
        columnTPM = columnHits+2
        columnTPHR = columnHits+4
        columnFAR = columnHits+5
        columnDiscrimination = columnHits+6
        columnC = columnHits+7
        rowToStart = rowHits
        wb = openpyxl.load_workbook(wbT, data_only=True)
        sheet = wb.get_sheet_by_name(sT)
        numAdjusted=0

        wbF = openpyxl.load_workbook(wbF, data_only=True)
        sheetF = wbF.get_sheet_by_name(sF)

        #Hits
        for y in range(rowToStart, 91, 16):
            sheet.cell(column = columnTPHits, row = observer+2,
                       value = sheetF.cell(column = columnHitsToTake, row = y).value)
            print "Hits: ", sheetF.cell(column=columnHitsToTake, row=y).value

            columnTPHits = columnTPHits + 10

        #False Alarms
        for y in range(rowToStart, 91, 16):
            sheet.cell(column = columnTPFA, row = observer+2,
                       value = sheetF.cell(column = columnHitsToTake+1, row = y).value)
            print "False Alarms: ", sheetF.cell(column = columnHitsToTake+1, row = y).value

            columnTPFA = columnTPFA + 10

        #Misses
        for y in range(rowToStart, 91, 16):
            sheet.cell(column = columnTPM, row = observer+2,
                       value = sheetF.cell(column = columnHitsToTake+2, row = y).value)


            columnTPM = columnTPM + 10

        #Hitrate
        for y in range(rowToStart+2, 94, 16):
            sheet.cell(column=columnTPHR, row=observer + 2, value=sheetF.cell(column=columnHitsToTake, row=y).value)
            if(sheetF.cell(column=columnHitsToTake, row=y).value == 0
                    or sheetF.cell(column=columnHitsToTake, row=y).value == 1):
                sheet.cell(column=columnTPHR, row=observer + 2,
                           value=sheetF.cell(column=columnHitsToTake, row=y+3).value)
                sheet.cell(column=columnTPHR + 4, row=observer + 2, value="Y")
                sheet.cell(column=columnTPHR + 5, row = observer + 2, value = "HR")
                numAdjusted = numAdjusted+1
                print "Hitrate numAdjusted: ", numAdjusted


            columnTPHR = columnTPHR + 10

            sheet.cell(column=columnFAR, row=observer + 2, value=sheetF.cell(column=columnHitsToTake + 1, row=y).value)
            if (sheetF.cell(column=columnHitsToTake + 1, row=y).value == 0 or sheetF.cell(column=columnHitsToTake + 1,
                                                                                          row=y).value == 1):
                sheet.cell(column=columnFAR, row=observer + 2,
                           value=sheetF.cell(column=columnHitsToTake, row=y + 4).value)
                numAdjusted = numAdjusted + 1

                sheet.cell(column=columnFAR + 3, row=observer + 2, value="Y")
                if (numAdjusted==2):
                    sheet.cell(column=columnFAR + 4, row=observer + 2, value="HR & FA")
                else:
                    sheet.cell(column=columnFAR + 4, row=observer + 2, value="FA")

            columnFAR = columnFAR + 10
            print numAdjusted
            numAdjusted = 0

        #FalseAlarmRate




        #Discrimination
        for y in range(rowToStart+10, 120, 16):
            sheet.cell(column=columnDiscrimination, row=observer + 2,
                       value=sheetF.cell(column=columnHitsToTake, row=y).value)


            columnDiscrimination= columnDiscrimination+10

        #C
        for y in range(rowToStart+11, 120, 16):
            sheet.cell(column=columnC, row=observer + 2,
                       value=sheetF.cell(column=columnHitsToTake, row=y).value)


            columnC = columnC + 10

        wb.save(fileName)